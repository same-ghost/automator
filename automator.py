# IMPORT ----------------------------------------------------------------------
import openpyxl
import time


# GLOBAL ----------------------------------------------------------------------
FILE_LOCATION = '/mnt/c/Users/talbert/Downloads/linux/'


START_COLUMN = 10


GER_COLUMN = {
'FYW': START_COLUMN, 
'WR' : START_COLUMN + 3, 
'FYS': START_COLUMN + 3, 
'NWL': START_COLUMN + 6, 
'NW' : START_COLUMN + 9, 
'HB' : START_COLUMN + 12, 
'TA' : START_COLUMN + 18, 
'HA' : START_COLUMN + 21, 
'VP' : START_COLUMN + 24, 
'MR' : START_COLUMN + 27, 
'FL' : START_COLUMN + 30, 
'UQ' : START_COLUMN + 33, 
'MB' : START_COLUMN + 36, 
'NE' : START_COLUMN + 39, 
'WC' : START_COLUMN + 42
}


# FUNCTION --------------------------------------------------------------------
def source_sheet():
    '''
        Open course listings workbook and return the active sheet.
    '''
    wb = openpyxl.load_workbook(FILE_LOCATION + 'datasource.xlsx')
    return wb['Degree Checklist - Course List ']
# END source_sheet()


def target_sheet():
    '''
        Open student listings workbook and return the active sheet.
    '''
    wb = openpyxl.load_workbook(FILE_LOCATION + 'skeleton.xlsx')
    return wb['Automator Student List']
# END target_sheet()


def validate_grade(grade):
    '''
        Check a single course grade against the list of valid grades.
    '''
    valid_grades = ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 
    'D+', 'D', 'D-', None]
    return grade in valid_grades
# END validator()


def validate_credit(credit_type):
    '''
        Process different credit types.
    '''
    pass
# END validate_credit()


def format_term(course_term):
    '''
        Returns the term a course was taken in the format 
        FA ##/SP ##/SU ##.
    '''
    terms = {'D1': 'FA', 'D2': 'SP', 'D4': 'SU'}
    if course_term != None:
        term = terms[course_term[4:6]] + ' ' + course_term[2:4]
        return term
    
    return None
# END format_term()


def ger(student, ger, course, term, grade):
    '''
        Scan a GER type, and apply it to the correct column.
    '''
    
    student[GER_COLUMN[ger]].value = course
    student[GER_COLUMN[ger] + 1].value = term
    student[GER_COLUMN[ger] + 2].value = grade
# END ger()


# MAIN ------------------------------------------------------------------------
source_sheet = source_sheet()
target_book = openpyxl.load_workbook(FILE_LOCATION + 'skeleton.xlsx')
target_sheet = target_book['Automator Student List']

print("Test begins: ")
t0 = time.process_time()

# Iterate through all students in Automator Student List
for student in target_sheet.iter_rows(min_row=2):
    target_id = student[0].value

    # Iterate through all courses in Automator Course List
    for course in source_sheet.iter_rows(min_row=2):
        
        # Validate grade (i.e., don't waste time on fails or PNP)
        course_grade = course[5].value
        if validate_grade(course_grade):

            # Make values human-parsable
            course_id = course[0].value
            course_code = course[2].value
            course_term = format_term(course[1].value)
            course_ger = course[6].value

            # If student and course match, and the course fulfills a GER,
            # pass the information along to the GER column sorter.
            if target_id == course[0].value and course_ger != None:
                ger(student, course_ger, course_code, course_term, course_grade)

# Print test results
t1 = time.process_time()
print(f"Test completed in {t1-t0} seconds.")

# Save workbook
target_book.save(FILE_LOCATION + 'targeted.xlsx')